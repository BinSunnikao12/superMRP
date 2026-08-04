import copy
import datetime
import openpyxl
import os
import pandas as pd
from pytz import timezone

from compute import compute
from data import data_mysql

vn_titles={
    "预计开工日期": "NGÀY BẮT ĐẦU DỰ KIẾN SẢN XUẤT",
    "预计完工日期": "NGÀY DỰ KIẾN HOÀN THÀNH",
    "客户订单号": "MÃ ĐƠN ĐẶT HÀNG CỦA KHÁCH HÀNG",
    "成品工单号": "MÃ CÔNG ĐƠN THÀNH PHẨM",
    "成品工单状态": "TRẠNG THÁI CÔNG ĐƠN THÀNH PHẨM",
    "父跟单码": "MÃ THEO ĐƠN XƯỞNG LẮP RÁP",
    "跟单码": "MÃ THEO ĐƠN",
    "主件": "LINH KIỆN CHỦ YẾU",
    "主件料名": "TÊN LIỆU CỦA LINH KIỆN CHÍNH",
    "主件规格": "QUY CÁCH LINH KIỆN CHÍNH",
    "主件成本中心": "TRUNG TÂM GIÁ THÀNH  LINH KIỆN CHÍNH",
    "下阶": "GIAI ĐOẠN SAU",
    "下阶料名": "TÊN LIỆU GIAI ĐOẠN SAU",
    "下阶规格": "QUY CÁCH GIAI ĐOẠN SAU",
    "下阶成本中心": "TRUNG TÂM GIÁ THÀNH GIAI ĐOẠN SAU",
    "BOM用量": "LƯỢNG DÙNG BOM",
    "毛需求": "TỔNG NHU CẦU",
    "总库存": "TỔNG TỒN KHO",
    "总在制": "TỔNG BÁN THÀNH PHẨM",
    "总装发出数量": "SỐ LƯỢNG PHÁT RA LẮP RÁP HOÀN CHỈNH",
    "GD01工单数": "SỐ CÔNG ĐƠN GD01",
    "总采购在途": "TỔNG MUA SẮM TRONG QUÁ TRÌNH VẬN CHUYỂN",
    "总工单供给": "CUNG CẤP TỔNG CÔNG ĐƠN",
    "总在验": " TỔNG SỐ ĐANG KIỂM NGHIỆM",
    "可用库存": "TỒN KHO CÓ THỂ DÙNG",
    "可用在制GD01": "BÁN THÀNH PHẨM CÓ THỂ DÙNG GD01",
    "可用工单数": "SỐ CÔNG ĐƠN CÓ THỂ DÙNG",
    "净需求": "NHU CẦU TỊNH",
    "合计欠料": "TỔNG SỐ THIẾU LIỆU",
    "子件标准人工工时": "THỜI GIAN TIÊU CHUẨN LINH KIỆN CON",
    "主件标准人工工时": "THỜI GIAN TIÊU CHUẨN LINH KIỆN CHÍNH",
    "补给策略": "CHÍNH SÁCH BÙ",
    "计划员": "NHÂN VIÊN KẾ HOẠCH",
    "采购物控人员": "NHÂN VIÊN KHỐNG CHẾ LIỆU THU MUA",
    "最近采购供应商": "NHÀ CUNG CẤP MUA HÀNG GẦN ĐÂY",
    "外购类型": "LOẠI HÌNH MUA NGOÀI",
    "材质": "CHẤT LIỆU",
    "最近采购供应商编码": "MÃ NHÀ CUNG CẤP MUA HÀNG GẦN ĐÂY",
    "采购单位批量": "ĐƠN VỊ MUA HÀNG SỐ LƯỢNG LỚN",
    "最小采购数量": "LƯỢNG MUA HÀNG NHỎ NHẤT",
    "采购单位": "ĐƠN VỊ MUA HÀNG",
    "生产单位批量": "ĐƠN VỊ SẢN XUẤT SỐ LƯỢNG LỚN",
    "最小生产数量": "SỐ LƯỢNG SẢN XUẤT NHỎ NHẤT",
    "生产单位": "ĐƠN VỊ SẢN XUẤT",
    "生产损耗率": "TỈ LỆ HƯ HẠI SẢN XUẤT",
    "固定生产前置时间": "THIẾT LẬP THỜI GIAN CỐ ĐỊNH TRƯỚC SẢN XUẤT",
    "变动生产前置时间": "THIẾT LẬP THỜI GIAN BIẾN ĐỘNG TRƯỚC SẢN XUẤT",
    "QC前置时间": "THIẾT LẬP THỜI GIAN TRƯỚC QC",
    "累计前置时间": "TỔNG CỘNG THỜI GIAN THIẾT LẬP",
    "来源单号": "NGUỒN GỐC MÃ ĐƠN",
    "需求计算方式": "PHƯƠNG THỨC TÍNH TOÁN YẾU CẦU",
    "齐套数": "SỐ ĐỦ BỘ",
    "主件单位": "ĐƠN VỊ LINH KIỆN CHÍNH",
    "下阶单位": "ĐƠN VỊ GIAI ĐOẠN SAU",
    "总装锁定日期": "CHỐT THỜI GIAN TỔNG LẮP RÁP",
    "立柱锁定日期": "CHỐT THỜI GIAN LẮP RÁP CHÂN",
    "承诺交期": "THỜI GIAN CAM KẾT GIAO HÀNG",
    "采购回复": "TRẢ LỜI THU MUA",
    "总装车间": "XƯỞNG LẮP RÁP",
    "越南主件料名": " TÊN LIỆU LINH KIỆN CHÍNH VIỆT NAM",
    "越南主件规格": "QUI CÁCH LIỆU CHÍNH VIỆT NAM",
    "越南下阶料名": "TÊN LIỆU VIỆT NAM GIAI ĐOẠN SAU",
    "越南下阶规格": "QUY CÁCH GIAI ĐOẠN SAU VIỆT NAM",
    "料号": "MÃ LIỆU",
    "料名": "TÊN LIỆU",
    "规格": "QUI CÁCH",
    "多余库存": "TỒN KHO DƯ",
    "多余在制": "BÁN THÀNH PHẨM DƯ",
    "安全库存量": "LƯỢNG TỒN KHO AN TOÀN",
    "总需求": "TỔNG YÊU CẦU",
    "多余库存含在途": "BAO GỒM LƯỢNG TỒN KHO TRÊN ĐƯỜNG VẬN CHUYỂN",
    "要求到货日期": "YÊU CẦU THỜI GIAN GIAO HÀNG",
    "需求": "YÊU CẦU",
    "采购单": "ĐƠN THU MUA",
    "单据日期": "NGÀY DỮ LIỆU CỦA ĐƠN",
    "采购单要求交期": "YÊU CẦU GIAO HÀNG ĐƠN THU MUA",
    "请购创建日期": "NGÀY TẠO YÊU CẦU MUA",
    "供应商": "NHÀ CUNG CẤP",
    "成本中心": "TRUNG TÂM GIÁ THÀNH",
    "是否新品": "CÓ PHẢI LIỆU MỚI",
    "多角贸易": "GIAO DICH ĐA GỐC",
    "接单日期": "NGÀY NHẬN ĐƠN",
    "约定交货日期": "HẸN THỜI GIAN GIAO HÀNG",
    "销售订单号": "MÃ ĐƠN HÀNG  KINH DOANH",
    "应发数量": "SỐ LƯỢNG PHẢI PHÁT",
    "缺料数量": "SỐ LƯỢNG THIẾU LIỆU",
    "备注": "GHI CHÚ",
    "净需求总和": "TỔNG SỐ NHU CẦU TỊNH",
    "齐套数总和": "TỔNG SỐ LƯỢNG ĐỦ BỘ",
    "人工工时": "THỜI GIAN NHÂN CÔNG",
    "工单供给": "CUNG CẤP CÔNG ĐƠN",
    "工单单号": "MÃ CÔNG ĐƠN",
    "工单状态": "TRẠNG THÁI CÔNG ĐƠN",
    "主件料号": "MÃ LIỆU LINH KIỆN CHÍNH",
    "主件需求数量": "SỐ LƯỢNG YÊU CẦU LINH KIỆN CHÍNH",
    "业务人员": "NHÂN VIÊN NGHIỆP VỤ",
    "客户": "KHÁCH HÀNG",
}

def t():
    current_directory = os.getcwd()
    print(current_directory)

format_float=lambda value: "{:.3f}".format(value) if  value % 1 != 0 else  str(int(value))  # 如果小数部分不为0，则保留3位小数,否则，显示整数部分

def allocate (i,q,f):
    c = f[i["要求到货日期"]+i["下阶"]]
    first_key = next(iter(c.keys()), "")
    if first_key == "":
        i["成本中心"] += " " +('' if i["成本中心"]=='' else format_float(q))
    elif c[first_key] > q:
        c[first_key] -= q
        i["成本中心"] += first_key + ('' if i["成本中心"]=='' else format_float(q))
    elif c[first_key] == q:
        i["成本中心"] += first_key + ('' if i["成本中心"]=='' else format_float(q))
        del c[first_key]
    else:
        q -= c[first_key]
        i["成本中心"] += first_key + format_float(c[first_key]) + ","
        del c[first_key]
        allocate(i, q,f)

def writer_sheet(data, titles, worksheet):
    for i, j in enumerate(titles):
        worksheet.cell(1, i + 1, j)
        worksheet.cell(2, i + 1, vn_titles.get(j))
    for i1, j1 in enumerate(data):
        for i2, j2 in enumerate(titles):
            worksheet.cell(i1 + 3, i2 + 1, j1.get(j2))

def djmy(d):
    d["多角贸易"] = ''
    elements = d["跟单码"].split(',')
    for element in elements:
        if "." not in element and 'XS07' in element and ' 发越南' not in d["多角贸易"]:
            d["多角贸易"] += ' 发越南'
        elif "." not in element and 'XS09' in element and ' 发姜山' not in d["多角贸易"]:
            d["多角贸易"] += " 发姜山"
        elif "." not in element and 'XS79' in element and ' 发广西' not in d["多角贸易"]:
            d["多角贸易"] += " 发广西"

def f(site):
    ser =   compute.serve(site)
    ser.calculate()

    k_list = ['预计开工日期', '预计完工日期', '客户订单号', '成品工单号', '成品工单状态', '父跟单码', '跟单码', '主件', '主件料名', '主件规格', '主件成本中心', '下阶',
              '下阶料名', '下阶规格', '下阶成本中心', 'BOM用量', '毛需求', '总库存', '总在制',"总装发出数量", 'GD01工单数', "总采购在途", '总工单供给', '总在验', '可用库存',
              '可用在制', "GD01可用工单数",'净需求', '合计欠料','子件标准人工工时','主件标准人工工时', '补给策略', '计划员', '采购物控人员', '最近采购供应商',
              '外购类型','材质','最近采购供应商编码', '采购单位批量', '最小采购数量', '采购单位',
              '生产单位批量', '最小生产数量','生产单位', '生产损耗率', '固定生产前置时间', '变动生产前置时间',
              'QC前置时间', '累计前置时间', '来源单号', '需求计算方式', '齐套数',"主件单位","下阶单位","总装锁定日期","立柱锁定日期","承诺交期","采购回复","总装车间"]
    if site == 'YN':
        k_list += ["越南主件料名", "越南主件规格", "越南下阶料名", "越南下阶规格"]
    workbook = openpyxl.Workbook()

    r_list = ["料号", "料名", "规格", "总库存", "总在制", "多余库存", "多余在制", "安全库存量","总需求","总在验","总采购在途","多余库存含在途"]
    worksheet1 = workbook.create_sheet('库存情况', 1)
    writer_sheet(ser.r, r_list, worksheet1)

    wg_copy = None
    zj_data = None
    if len(ser.l) > 0:
        df = pd.DataFrame(ser.l)
        df_xs = df[df['成品工单号'].str.contains('XS') & ~df['跟单码'].str.contains('\.')]
        df_xs = df_xs.rename(columns={'预计开工日期': '约定交货日期', '成品工单号': '销售订单号', '下阶': '料号','下阶料名':"料名",
                                          "下阶规格":"规格","下阶成本中心":"成本中心","毛需求":"应发数量","净需求":"缺料数量"})
        ws = workbook.create_sheet('销售备货',3)
        df_xs['接单日期'] =  df['成品工单号'].apply(lambda x:ser.docdt.get(x).strftime("%Y-%m-%d") if ser.docdt.get(x) is not None else "")
        writer_sheet(df_xs.to_dict('records'),
                     ['接单日期','约定交货日期', '销售订单号', '料号', '料名', '规格', '成本中心', '应发数量', '缺料数量','总库存','总在制','总采购在途','总在验','备注'], ws)

        df_wg = df[df['补给策略'] == '外购']
        df_wg = df_wg[df_wg['净需求'] != 0]
        df_wg['下阶料名'] = df_wg['下阶料名'].fillna("")
        df_wg['下阶规格'] = df_wg['下阶规格'].fillna("")
        df_wg['采购物控人员'] = df_wg['采购物控人员'].fillna("")
        df_wg['最近采购供应商'] = df_wg['最近采购供应商'].fillna("")
        # grouped_wg = df_wg.groupby(
        #     ['采购物控人员', '预计开工日期', '最近采购供应商', '下阶', '下阶料名', '下阶规格', '总采购在途', '总库存', '总在制', '总在验', '外购类型']).apply(
        #         lambda x: pd.Series({
        #             '净需求': x['净需求'].sum(),
        #             '跟单码': ','.join(x['跟单码']),
        #             '总装锁定日期': ','.join(x['总装锁定日期']) if x['总装锁定日期'].any() else '',
        #             '立柱锁定日期': ','.join(x['立柱锁定日期']) if x['立柱锁定日期'].any() else '',
        #             '客户订单号': ','.join(x['客户订单号']),
        #             '主件成本中心': {cc: x[x['主件成本中心'] == cc]['净需求'].sum() for cc in x['主件成本中心'].unique() if pd.notnull(cc)}
        #         })
        #     ).reset_index()
        grouped_wg=df_wg
        grouped_wg = grouped_wg.rename(columns={'净需求': '净需求总和', '预计开工日期': '要求到货日期'})
        grouped_wg = grouped_wg.sort_values(by='要求到货日期')
        # grouped_wg["要求上线日期"] = grouped_wg["要求到货日期"]
        # 检查日期列的类型
        if pd.api.types.is_datetime64_any_dtype(grouped_wg['要求到货日期']):
            grouped_wg['要求到货日期'] = (pd.to_datetime(grouped_wg['要求到货日期']) - pd.DateOffset(days=1)).dt.strftime('%Y-%m-%d')
        else:
            grouped_wg['要求到货日期'] = (pd.to_datetime(grouped_wg['要求到货日期']) - pd.DateOffset(days=1)).dt.strftime('%Y-%m-%d')
        wg = grouped_wg.to_dict('records')
        wg_copy = copy.deepcopy(wg)
        wg_copy2 = copy.deepcopy(wg)
        f={}
        for i in wg_copy2:
            f[i["要求到货日期"]+i["下阶"]]=i["主件成本中心"]
        # writer_sheet(wg, ['采购物控人员','最近采购供应商','要求到货日期', '下阶','下阶料名','下阶规格','净需求总和','总采购在途','总库存','总在制','总在验','跟单码'], workbook.create_sheet("采购需求", 2))
        p = ser.purchase_order_detail_dict
        wg_d = []
        for i in wg_copy2:
            pp = p.get(i["下阶"])
            # del i["最近采购供应商"]
            x = i.pop("净需求总和")
            if pp:
                for j in pp:
                    if j["ZTNUM"] >= x:
                        ii = {**i, "采购单": j["CGD"], "单据日期": j["PMDLDOCDT"].strftime("%Y-%m-%d"),
                              "采购单要求交期": j["PMDO013"].strftime("%Y-%m-%d") if j["PMDO013"]!=None else None,
                              "请购创建日期": j["CJRQ"],
                              "供应商": j["PMAAL003"] if j["PMAAL003"]!=None else i["最近采购供应商"],
                              "需求": x,'未处理请购数':None,"采购物控人员":j["OOAG011"] if j["OOAG011"]!=None else i["采购物控人员"],
                              "公司型号":j["公司型号"],"客户型号":j["客户型号"]}

                        if 'QG34' in ii["采购单"] and ser.site=='LG' :
                            ii["采购物控人员"] = "千银海"
                        j["ZTNUM"] -= x
                        x = 0
                        wg_d.append(ii)
                        break
                    elif j["ZTNUM"] == 0:
                        continue
                    else:
                        ii = {**i, "采购单": j["CGD"], "单据日期": j["PMDLDOCDT"].strftime("%Y-%m-%d"),
                              "采购单要求交期":j["PMDO013"].strftime("%Y-%m-%d") if j["PMDO013"]!=None else None,
                              "请购创建日期": j["CJRQ"],
                              "供应商":  j["PMAAL003"] if j["PMAAL003"]!=None else i["最近采购供应商"],
                              "需求": j["ZTNUM"],'未处理请购数':None,"采购物控人员":j["OOAG011"] if j["OOAG011"]!=None else i["采购物控人员"],
                              "公司型号":j["公司型号"],"客户型号":j["客户型号"]}
                        if 'QG34' in ii["采购单"] and ser.site=='LG' :
                            ii["采购物控人员"] = "千银海"
                        x = x - j["ZTNUM"]
                        j["ZTNUM"] = 0
                        wg_d.append(ii)
            # if x > 0:
            #     qg_qty=ser.qg_map.get(i["下阶"],0)
            #     q=0
            #     if x<qg_qty:
            #         q=x
            #         ser.qg_map[i["下阶"]]=qg_qty-x
            #     else:
            #         q=qg_qty
            #         ser.qg_map[i["下阶"]]=0
            #     ii = {**i, "采购单": None, "单据日期": None, "采购单要求交期": None, "请购创建日期": None, "供应商": None, "需求": x,'未处理请购数':q,"公司型号":None,"客户型号":None}
            #     wg_d.append(ii)
        # up=ser.unaudited_purchase_order_dict

        def cg_xp(x):
            base = ser.base_dict[x]
            if base["IMAFUD003"] != "Y" and base["IMAFUA003"] == "Y":
                return "新品已试产"
            elif base["IMAFUD003"] != "Y":
                return "新品"
            else:
                return ""

        for i in wg_d:
            remark = ser.remark.get(i["跟单码"])
            i["采购回复"] = remark[1] if remark != None else ""
            i["承诺交期"] = remark[2] if remark != None else ""
            i["是否新品"]=cg_xp(i["下阶"])
            i["成本中心"] = i["主件成本中心"]
            # allocate(i, i["需求"],f)
            djmy(i)  #判断是不是要发往其他基地

            i["采购物控人员"]=ser.ryjj.get(i["采购物控人员"],i["采购物控人员"])
            for jjj in ser.l:
                if i["跟单码"]==jjj["跟单码"]:
                    jjj["采购物控人员"]=i["采购物控人员"]
            if ser.site == "YN" or ser.site == "FN":
                vn_items_f = ser.vn_items_dict.get(i["下阶"])
                i["越南下阶料名"] = vn_items_f.get("IMAAL003", '') if vn_items_f else ''
                i["越南下阶规格"] = vn_items_f.get("IMAAL004", '') if vn_items_f else ''
        cg_titles=['采购物控人员', '要求到货日期', '下阶', '下阶料名', '下阶规格', '需求', '总采购在途', '总库存', '总在制', '总在验',
                            "采购单", "单据日期","采购单要求交期",  "请购创建日期", "供应商", '外购类型', '客户订单号', '跟单码','成本中心','总装锁定日期','立柱锁定日期',
                            '是否新品','多角贸易','承诺交期','采购回复']
        cg_sheetname='采购需求-明细'
        if ser.site == "YN" or ser.site == "FN":
            cg_sheetname+='chi tiết'
            cg_titles += ["越南下阶料名", "越南下阶规格"]
        writer_sheet(wg_d, cg_titles, workbook.create_sheet(cg_sheetname, 2))

        worksheet = workbook.create_sheet('物料需求', 0)
        writer_sheet(ser.l, k_list, worksheet)
        # cgzt=[]
        # for i in ser.purchase_order_detail_dict.values():
        #         cgzt+=i
        # for i in cgzt:
        #     i["料名"]=ser.items_dict[i["PMDO001"]]["IMAAL003"]
        #     i["规格"]=ser.items_dict[i["PMDO001"]]["IMAAL004"]
        #     i["外购类型"] = ser.outsourcing_type_dict.get(i["PMDO001"])
        #     i["料号"]=i.pop("PMDO001")
        #     i["无单采购数量"]=i.pop("ZTNUM")
        #     i["采购数量"]=i.pop("ZTNUM2")
        #     i["供应商编码"]=i.pop("PMDL004")
        #     i["供应商"] = i.pop("PMAAL003")
        #     i["采购单"] =i.pop("CGD")
        #     i["采购下单日期"]=i.pop("PMDLDOCDT")
        #     i["采购要求交期"]=i.pop("PMDO013")
        #     i["请购创建日期"]=i.pop("CJRQ")
        # writer_sheet(cgzt, ["采购单","料号","料名","规格","外购类型","无单采购数量","采购数量","供应商编码","供应商","采购下单日期","采购要求交期","请购创建日期"],
        #              workbook.create_sheet("采购订单", 3))

        df_zz = df[df['补给策略'] != '外购']
        df_zz = df_zz[df_zz['净需求'] != 0]
        df_zz['下阶料名'] = df_zz['下阶料名'].fillna("")
        df_zz['下阶规格'] = df_zz['下阶规格'].fillna("")
        df_zz['下阶成本中心'] = df_zz['下阶成本中心'].fillna("")
        titles = ['下阶成本中心', '预计完工日期', '下阶', '下阶料名', '下阶规格', '净需求总和', '齐套数总和', '人工工时',
                  '工单供给', '总工单供给', '跟单码','总装锁定日期','立柱锁定日期',"是否新品","多角贸易","主件成本中心"]
        df_zz['喷涂料名'] = df_zz['主件料名']
        df_zz.loc[(~(df_zz['下阶成本中心'].isin(['焊接车间-乐歌', '激光切割车间']))) | (df_zz['主件成本中心'] != '喷塑车间-乐歌'), '喷涂料名'] = ''
        grouped = df_zz.groupby(['下阶成本中心','下阶成本中心编码', '预计开工日期', '下阶', '下阶料名', '下阶规格', '总工单供给','喷涂料名',"主件成本中心"]).agg(
            {'净需求': 'sum', '齐套数': 'sum', '人工工时': "sum", '跟单码': lambda x: ','.join(x),
             '总装锁定日期':  lambda x: ','.join(x) if x.any() else '',
             '立柱锁定日期': lambda x: ','.join(x) if x.any() else ''
             }).reset_index()
        grouped = grouped.rename(columns={'净需求': '净需求总和', '齐套数': '齐套数总和', '预计开工日期': '预计完工日期'})
        # grouped["是否新品"]=df['下阶'].apply(xp)
        def zj_xp(x):
            base = ser.base_dict[x]
            if base["IMAFUD004"] != "Y" and base["IMAFUA004"] == "Y":
                return "新品已试产"
            elif base["IMAFUD004"] != "Y":
                return "新品"
            else:
                return ""
        # 检查日期列的类型
        if pd.api.types.is_datetime64_any_dtype(grouped['预计完工日期']):
            grouped['预计完工日期'] = (grouped['预计完工日期'] - pd.DateOffset(days=1)).dt.strftime('%Y-%m-%d')
        else:
            grouped['预计完工日期'] = (pd.to_datetime(grouped['预计完工日期']) - pd.DateOffset(days=1)).dt.strftime('%Y-%m-%d')
        zgd = copy.copy(ser.production_supply_dict)
        for index, row in grouped.iterrows():  # 计算单行的工单供给
            if zgd.get(row['下阶'], 0) > row['净需求总和']:
                grouped.loc[index, '工单供给'] = row['净需求总和']  # 在A中加入一列，并赋值为qty
                zgd[row['下阶']] -= row['净需求总和']  # 将B的value减去qty
            else:
                grouped.loc[index, '工单供给'] = zgd.get(row['下阶'], 0)  # 在A中加入一列，并赋值为B的value
                zgd[row['下阶']] = 0  # 将B的value设为0
        zj_data = grouped.to_dict('records')
        for i in zj_data:
            i["是否新品"]=zj_xp(i["下阶"])
            djmy(i)  # 判断是不是要发往其他基地
            if ser.site == "YN" or ser.site == "FN":
                vn_items_f = ser.vn_items_dict.get(i["下阶"])
                i["越南下阶料名"] = vn_items_f.get("IMAAL003", '') if vn_items_f else ''
                i["越南下阶规格"] = vn_items_f.get("IMAAL004", '') if vn_items_f else ''
        if ser.site == "YN" or ser.site == "FN":
            titles+=["越南下阶料名","越南下阶规格"]
        writer_sheet(zj_data, titles, workbook.create_sheet("生产需求"))
        titles.pop(0)  # 去掉下阶成本中心
        title_list=(df_zz['下阶成本中心'].unique())
        sort_dict={
            'LG':['委外需求','激光切割车间','焊接车间-乐歌','喷塑车间-乐歌','总装一车间','总装二车间','总装三车间','总装四车间','木工车间']
        }
        list_sort=sort_dict.get(site,[])
        title_list=sorted(title_list, key=lambda x: list_sort.index(x) if x in list_sort else len(list_sort))
        for i, a_value in enumerate(title_list):
            titles1=copy.copy(titles)
            if a_value in ['焊接车间-乐歌','激光切割车间']:
                titles1.append("喷涂料名")
            ws = workbook.create_sheet(a_value)
            a_data = grouped[grouped['下阶成本中心'] == a_value].reset_index(drop=True)
            d=a_data.to_dict('records')
            for i in d:
                i["是否新品"] = zj_xp(i["下阶"])
                djmy(i)  # 判断是不是要发往其他基地
                if ser.site == "YN" or ser.site == "FN":
                    vn_items_f = ser.vn_items_dict.get(i["下阶"])
                    i["越南下阶料名"] = vn_items_f.get("IMAAL003", '') if vn_items_f else ''
                    i["越南下阶规格"] = vn_items_f.get("IMAAL004", '') if vn_items_f else ''
            writer_sheet(d, titles1, ws)

        # 从列表包字典的结构中取出需要的列
        n = pd.DataFrame(ser.need,columns=['预计开工日期', '预计完工日期', '主件料号', '主件需求数量', '工单状态', '工单单号','SFBASEQ', 'SFAAUD002', 'OOAG011','客户'])
        n = n.rename(columns={'SFAAUD002': '客户订单号', "OOAG011": "业务人员"})  # 重命名列名为 '客户订单号'
        n['预计开工日期'] = n['预计开工日期'].dt.strftime('%Y-%m-%d')
        n['预计完工日期'] = n['预计完工日期'].dt.strftime('%Y-%m-%d')
        n[['主件料名', '主件规格']] = n['主件料号'].apply(
            lambda x: pd.Series([ser.items_dict[x]['IMAAL003'], ser.items_dict[x]['IMAAL004']]))
        def process_work_order(row):
            work_order = row['工单单号']
            if len(work_order) >= 5 and work_order[3:5] == 'GD':
                return 0
            return row['SFBASEQ']
        n['SFBASEQ'] = n.apply(process_work_order, axis=1)
        n = n.drop_duplicates()
        ws = workbook.create_sheet('主计划')
        writer_sheet(n.to_dict('records'),
                     ['预计开工日期', '预计完工日期', '客户订单号', '工单单号','工单状态',
                      '主件料号', '主件料名','主件规格', '主件需求数量', "业务人员","客户"], ws)

        sheet_names = ['Sheet']  # 删除空白页
        for i in sheet_names:
            if i in workbook.sheetnames:
                workbook.remove(workbook[i])

    version = site + "_" + datetime.datetime.now(timezone('Asia/Shanghai')).strftime("%y%m%d%H%M%S")

    current_directory = os.path.dirname(__file__)
    parent_directory =os.path.dirname(os.path.dirname(current_directory))
    file_path = os.path.join(parent_directory, site)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    workbook.save(os.path.join(file_path, '{}.xlsx'.format(version)))
    workbook.close()

    data_mysql.mysql_load(ser.l, site, version, wg_d, zj_data)

    file_names = os.listdir(file_path)
    max_file_name = max(file_names)
    site_elements = [element for element in file_names if element.startswith(max_file_name[:9])]

    # 删除当天最大文件之外的所有文件
    for file_name in site_elements:
        if file_name != max_file_name:
            os.remove(os.path.join(file_path, file_name))
